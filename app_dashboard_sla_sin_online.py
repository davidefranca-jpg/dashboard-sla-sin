import os
import sys
import csv
import io
import re
import json
import uuid
import html
import tempfile
import traceback
import pickle
import hashlib
import hmac
import secrets
from http.cookies import SimpleCookie
from datetime import datetime, date, timedelta
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs

try:
    import openpyxl
except Exception:
    print("ERRO: biblioteca openpyxl nao instalada.")
    print("Rode no PowerShell:")
    print(r"cd C:\IA\Projeto")
    print(r".\venv\Scripts\activate")
    print("pip install openpyxl")
    sys.exit(1)

HOST = "0.0.0.0"
PORT = int(os.environ.get("PORT", 8090))
APP_DIR = r"C:\IA\Projeto"
UPLOAD_DIR = os.path.join(APP_DIR, "uploads_sla") if os.name == "nt" else tempfile.gettempdir()
os.makedirs(UPLOAD_DIR, exist_ok=True)

# Pasta fixa para guardar a análise processada.
# Antes o dashboard guardava tudo apenas na memória (CACHE = {}).
# Quando o servidor reiniciava, dormia ou a plataforma online limpava a memória,
# o token expirava e a tela pedia para subir a planilha novamente.
# Agora cada análise fica salva em disco e pode ser consultada durante o dia todo.
CACHE_DIR = os.path.join(APP_DIR, "cache_sla") if os.name == "nt" else os.path.join(tempfile.gettempdir(), "cache_sla")
os.makedirs(CACHE_DIR, exist_ok=True)
CACHE = {}
CACHE_VALID_HOURS = int(os.environ.get("CACHE_VALID_HOURS", "36"))

USERS_FILE = os.path.join(APP_DIR, "usuarios_sla.json") if os.name == "nt" else os.path.join(tempfile.gettempdir(), "usuarios_sla.json")
SESSIONS = {}
SESSION_HOURS = int(os.environ.get("SESSION_HOURS", "12"))
ADMIN_LOGIN = os.environ.get("ADMIN_LOGIN", "admin")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "Admin@123")
DEFAULT_CLIENT_PASSWORD_SUFFIX = os.environ.get("DEFAULT_CLIENT_PASSWORD_SUFFIX", "@123")

def now_ts():
    return datetime.now().timestamp()

def hash_password(password, salt=None):
    if salt is None:
        salt = secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), 120000)
    return salt + "$" + digest.hex()

def check_password(password, stored_hash):
    try:
        salt, digest = stored_hash.split("$", 1)
        test_hash = hash_password(password, salt).split("$", 1)[1]
        return hmac.compare_digest(test_hash, digest)
    except Exception:
        return False

def load_users():
    if not os.path.exists(USERS_FILE):
        users = {
            ADMIN_LOGIN: {
                "senha_hash": hash_password(ADMIN_PASSWORD),
                "tipo": "admin",
                "nome": "Administrador",
                "cliente_codigo": "",
                "cliente_nome": "",
                "trocar_senha": True,
            }
        }
        save_users(users)
        return users
    try:
        with open(USERS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        traceback.print_exc()
        return {}

def save_users(users):
    tmp = USERS_FILE + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(users, f, ensure_ascii=False, indent=2)
    os.replace(tmp, USERS_FILE)

def user_safe(u):
    if not u:
        return None
    return {k: v for k, v in u.items() if k != "senha_hash"}

def get_session_user(handler):
    cookie_header = handler.headers.get("Cookie", "")
    if not cookie_header:
        return None
    cookie = SimpleCookie()
    try:
        cookie.load(cookie_header)
    except Exception:
        return None
    sid = cookie.get("sid")
    if not sid:
        return None
    sid = sid.value
    session = SESSIONS.get(sid)
    if not session:
        return None
    if now_ts() - session.get("created", 0) > SESSION_HOURS * 3600:
        SESSIONS.pop(sid, None)
        return None
    users = load_users()
    login = session.get("login")
    u = users.get(login)
    if not u:
        return None
    u = dict(u)
    u["login"] = login
    return u

def create_session(handler, login):
    sid = secrets.token_urlsafe(32)
    SESSIONS[sid] = {"login": login, "created": now_ts()}
    handler.send_header("Set-Cookie", f"sid={sid}; Path=/; HttpOnly; SameSite=Lax; Max-Age={SESSION_HOURS*3600}")

def clear_session(handler):
    handler.send_header("Set-Cookie", "sid=; Path=/; HttpOnly; SameSite=Lax; Max-Age=0")

def normalize_cliente_codigo(v):
    s = norm_text(v)
    if s.endswith(".0"):
        s = s[:-2]
    return s

def cliente_login_from_codigo(codigo):
    codigo = re.sub(r"[^A-Za-z0-9_.-]", "_", normalize_cliente_codigo(codigo))
    return f"cliente_{codigo}"

def auto_create_client_users(rows):
    """
    Cria usuários de cliente automaticamente a partir da aba Rastreamento:
    AF = código cliente
    AG = nome cliente

    Login gerado: cliente_<codigo>
    Senha inicial: <codigo>@123
    A senha fica criptografada no arquivo usuarios_sla.json.
    """
    users = load_users()
    created = []
    for r in rows:
        codigo = normalize_cliente_codigo(r.get("cliente_codigo", ""))
        nome = norm_text(r.get("cliente_nome", ""))
        if not codigo:
            continue
        login = cliente_login_from_codigo(codigo)
        if login in users:
            if nome and not users[login].get("cliente_nome"):
                users[login]["cliente_nome"] = nome
            continue
        senha_inicial = f"{codigo}{DEFAULT_CLIENT_PASSWORD_SUFFIX}"
        users[login] = {
            "senha_hash": hash_password(senha_inicial),
            "tipo": "cliente",
            "nome": nome or login,
            "cliente_codigo": codigo,
            "cliente_nome": nome,
            "trocar_senha": True,
        }
        created.append((login, senha_inicial, codigo, nome))
    if created:
        save_users(users)
    return created

def latest_file_path():
    return os.path.join(CACHE_DIR, "_latest_token.json")

def save_latest_token(token):
    with open(latest_file_path(), "w", encoding="utf-8") as f:
        json.dump({"token": token, "updated_at": datetime.now().isoformat()}, f, ensure_ascii=False, indent=2)

def load_latest_token():
    path = latest_file_path()
    if not os.path.exists(path):
        return ""
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f).get("token", "")
    except Exception:
        return ""

def filter_rows_for_user(rows, user):
    if not user:
        return []
    if user.get("tipo") == "admin":
        return rows
    codigo = normalize_cliente_codigo(user.get("cliente_codigo", ""))
    return [r for r in rows if normalize_cliente_codigo(r.get("cliente_codigo", "")) == codigo]

def require_login_html():
    return render_login("Faça login para acessar o sistema.")


MONTHS_PT = {
    1: "Jan", 2: "Fev", 3: "Mar", 4: "Abr", 5: "Mai", 6: "Jun",
    7: "Jul", 8: "Ago", 9: "Set", 10: "Out", 11: "Nov", 12: "Dez"
}

COLS = {
    "nf": 1, "data_inicial": 3, "uf_origem": 7, "tipo": 9, "entrega": 10,
    "link": 11, "destinatario": 13, "cidade_destino": 14, "uf_destino": 15,
    "data_entrega": 18, "ocorrencia": 19, "filial": 25, "sla": 28,
    "parceiro": 29, "sla_justificado": 30, "uf_parceiro": 31, "cliente_codigo": 32, "cliente_nome": 33
}

def norm_text(v):
    if v is None:
        return ""
    s = str(v).strip()
    return " ".join(s.split())

def norm_key(v):
    s = norm_text(v).upper()
    trans = str.maketrans("ÁÀÂÃÉÈÊÍÌÎÓÒÔÕÚÙÛÇ", "AAAAEEEIIIOOOOUUUC")
    return s.translate(trans)

def to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if v is None or str(v).strip() == "":
        return None
    if isinstance(v, (int, float)):
        try:
            return (date(1899, 12, 30) + timedelta(days=int(v)))
        except Exception:
            return None
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except Exception:
            pass
    return None

def to_number(v, default=0):
    if v is None or str(v).strip() == "":
        return default
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).replace(",", ".")
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    return int(float(m.group(0))) if m else default

def fmt_date(d):
    if not d:
        return ""
    return d.strftime("%d/%m/%Y")


def cache_file_path(token):
    safe = re.sub(r"[^A-Za-z0-9_.-]", "", token or "")
    if not safe:
        return ""
    return os.path.join(CACHE_DIR, f"{safe}.pkl")

def save_cache(token, rows):
    """Salva a análise em disco para não expirar ao atualizar a página."""
    payload = {
        "created_at": datetime.now(),
        "rows": rows,
    }
    path = cache_file_path(token)
    with open(path, "wb") as f:
        pickle.dump(payload, f, protocol=pickle.HIGHEST_PROTOCOL)

def load_cache(token):
    """Carrega análise da memória ou do disco, mantendo disponível durante o dia."""
    if token in CACHE:
        return CACHE[token]

    path = cache_file_path(token)
    if not path or not os.path.exists(path):
        return None

    try:
        with open(path, "rb") as f:
            payload = pickle.load(f)
        created_at = payload.get("created_at")
        rows = payload.get("rows")
        if not rows:
            return None

        if isinstance(created_at, datetime):
            idade = datetime.now() - created_at
            if idade > timedelta(hours=CACHE_VALID_HOURS):
                return None

        CACHE[token] = rows
        return rows
    except Exception:
        traceback.print_exc()
        return None


def is_business_day(d, holidays):
    return d.weekday() < 5 and d not in holidays

def add_business_days(start, days, holidays):
    if not start:
        return None
    days = max(0, int(days or 0))
    d = start
    added = 0
    while added < days:
        d += timedelta(days=1)
        if is_business_day(d, holidays):
            added += 1
    return d

def business_days_between(start, end, holidays):
    if not start or not end or end <= start:
        return 0
    d = start + timedelta(days=1)
    total = 0
    while d <= end:
        if is_business_day(d, holidays):
            total += 1
        d += timedelta(days=1)
    return total

def load_holidays(wb):
    holidays = set()
    if "Feriado - final de semana" in wb.sheetnames:
        ws = wb["Feriado - final de semana"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            for v in row[:3]:
                d = to_date(v)
                if d:
                    holidays.add(d)
    if "Listas" in wb.sheetnames:
        ws = wb["Listas"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = to_date(row[13] if len(row) > 13 else None)
            if d:
                holidays.add(d)
    return holidays

def load_ocorrencias_justificadas(wb):
    itens = set()
    if "ocorrencias" not in wb.sheetnames:
        return itens
    ws = wb["ocorrencias"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        for idx in (0, 6, 14):
            if idx < len(row):
                s = norm_key(row[idx])
                if s:
                    itens.add(s)
    return itens

def classify_row(r, holidays, occ_justificadas, hoje):
    start = r["data_inicial"]
    delivered = r["data_entrega"]
    due = add_business_days(start, r["sla"], holidays)
    entrega_key = norm_key(r["entrega"])

    # STATUS GERAL CORRIGIDO:
    # Agora o dashboard separa:
    # 1) Entregue antes do prazo  -> data de entrega menor que data prevista
    # 2) Entregue no prazo        -> data de entrega igual a data prevista
    # 3) Entregue atrasado        -> data de entrega maior que data prevista
    # Antes o script somava "antes do prazo" dentro de "Entregue no prazo",
    # por isso o Status geral ficava diferente do resumo da planilha.
    sj_key = norm_key(r.get("sla_justificado_raw"))
    justified_flag = sj_key == "SLA JUSTIFICADO" or "SLA JUSTIFICADO" in sj_key

    is_delivered = delivered is not None or "ENTREGUE" in entrega_key

    if justified_flag:
        status = "Entregue atrasado - Justificado"
        finalizado = True
    elif is_delivered and delivered and due and delivered < due:
        status = "Entregue antes do prazo"
        finalizado = True
    elif is_delivered and delivered and due and delivered == due:
        status = "Entregue no prazo"
        finalizado = True
    elif is_delivered and delivered and due and delivered > due:
        status = "Entregue atrasado"
        finalizado = True
    elif is_delivered:
        status = "Entregue atrasado"
        finalizado = True
    elif due and hoje <= due:
        status = "Em aberto no prazo"
        finalizado = False
    else:
        status = "Em aberto com atraso"
        finalizado = False

    # CORRECAO ABA RASTREAMENTO - COLUNA K (LINKS):
    # Considera Falta Link quando a coluna K for exatamente "0".
    # Nao depende de estar entregue/finalizado.
    link_val = norm_text(r["link"])
    falta_link = link_val == "0"

    dias_atraso = 0
    if status == "Em aberto com atraso" and due:
        dias_atraso = business_days_between(due, hoje, holidays)
    elif status in ("Entregue atrasado", "Entregue atrasado - Justificado") and due:
        fim = delivered if delivered else hoje
        dias_atraso = business_days_between(due, fim, holidays)

    return status, finalizado, due, dias_atraso, falta_link


def read_workbook(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if "Rastreamento" not in wb.sheetnames:
        raise ValueError('A planilha precisa ter a aba "Rastreamento".')
    holidays = load_holidays(wb)
    occ_justificadas = load_ocorrencias_justificadas(wb)
    ws = wb["Rastreamento"]
    hoje = date.today()
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        nf = norm_text(row[COLS["nf"]-1] if len(row) >= COLS["nf"] else None)
        if not nf:
            continue
        r = {
            "linha": i,
            "nf": nf,
            "data_inicial": to_date(row[COLS["data_inicial"]-1]),
            "uf_origem": norm_text(row[COLS["uf_origem"]-1]),
            "tipo": norm_text(row[COLS["tipo"]-1]),
            "entrega": norm_text(row[COLS["entrega"]-1]),
            "link": norm_text(row[COLS["link"]-1]),
            "destinatario": norm_text(row[COLS["destinatario"]-1]),
            "cidade_destino": norm_text(row[COLS["cidade_destino"]-1]),
            "uf_destino": norm_text(row[COLS["uf_destino"]-1]),
            "data_entrega": to_date(row[COLS["data_entrega"]-1]),
            "ocorrencia": norm_text(row[COLS["ocorrencia"]-1]),
            "filial": norm_text(row[COLS["filial"]-1]),
            "sla": to_number(row[COLS["sla"]-1], 0),
            "parceiro": norm_text(row[COLS["parceiro"]-1]),
            "uf_parceiro": norm_text(row[COLS["uf_parceiro"]-1]) if len(row) >= COLS["uf_parceiro"] else "",
            "cliente_codigo": normalize_cliente_codigo(row[COLS["cliente_codigo"]-1]) if len(row) >= COLS["cliente_codigo"] else "",
            "cliente_nome": norm_text(row[COLS["cliente_nome"]-1]) if len(row) >= COLS["cliente_nome"] else "",
            "sla_justificado_raw": row[COLS["sla_justificado"]-1],
        }
        r["rota_uf"] = f'{r["uf_origem"]} x {r["uf_destino"]}'.strip(" x")
        r["parceiro_uf"] = f'{r["parceiro"] or "Nao informado"} - {r["uf_parceiro"] or "UF nao informada"}'
        r["operacao_grupo"] = "Devolucao/Reversa" if any(x in norm_key(r["tipo"]) for x in ["DEVOL", "REVERS"]) else "Normal"
        status, finalizado, due, dias_atraso, falta_link = classify_row(r, holidays, occ_justificadas, hoje)
        r.update({"status": status, "finalizado": finalizado, "data_prevista": due, "dias_atraso": dias_atraso, "falta_link": falta_link})
        rows.append(r)
    return rows, holidays

def group_count(rows, key, only_status=None):
    d = {}
    for r in rows:
        if only_status and r["status"] not in only_status:
            continue
        k = r.get(key) or "Nao informado"
        d.setdefault(k, 0)
        d[k] += 1
    return sorted(d.items(), key=lambda x: x[1], reverse=True)

def pct(a, b):
    return 0 if not b else round((a / b) * 100, 2)

def make_summary(rows):
    total = len(rows)
    counts = {}
    for r in rows:
        counts[r["status"]] = counts.get(r["status"], 0) + 1
    finalizados = sum(1 for r in rows if r["finalizado"])
    em_aberto = total - finalizados
    no_prazo_real = counts.get("Entregue antes do prazo", 0) + counts.get("Entregue no prazo", 0)
    atrasado_real = counts.get("Entregue atrasado", 0) + counts.get("Entregue atrasado - Justificado", 0)
    no_prazo_just = counts.get("Entregue antes do prazo", 0) + counts.get("Entregue no prazo", 0) + counts.get("Entregue atrasado - Justificado", 0)
    atrasado_just = counts.get("Entregue atrasado", 0)
    falta_link = sum(1 for r in rows if r["falta_link"])
    return {
        "total": total,
        "finalizados": finalizados,
        "em_aberto": em_aberto,
        "pct_finalizados": pct(finalizados, total),
        "pct_em_aberto": pct(em_aberto, total),
        "counts": counts,
        "falta_link": falta_link,
        "sla_real_ok": pct(no_prazo_real, max(1, no_prazo_real + atrasado_real)),
        "sla_real_atrasado": pct(atrasado_real, max(1, no_prazo_real + atrasado_real)),
        "sla_just_ok": pct(no_prazo_just, max(1, no_prazo_just + atrasado_just)),
        "sla_just_atrasado": pct(atrasado_just, max(1, no_prazo_just + atrasado_just)),
    }

def csv_bytes(rows):
    out = io.StringIO()
    w = csv.writer(out, delimiter=';')
    w.writerow(["NF", "Código Cliente", "Nome Cliente", "Status", "Data inicial", "Data prevista", "Data entrega", "Dias atraso", "Parceiro", "Filial", "Tipo", "UF origem", "UF destino", "Destinatario", "Ocorrencia", "Falta link"])
    for r in rows:
        w.writerow([r["nf"], r.get("cliente_codigo", ""), r.get("cliente_nome", ""), r["status"], fmt_date(r["data_inicial"]), fmt_date(r["data_prevista"]), fmt_date(r["data_entrega"]), r["dias_atraso"], r["parceiro"], r["filial"], r["tipo"], r["uf_origem"], r["uf_destino"], r["destinatario"], r["ocorrencia"], "SIM" if r["falta_link"] else "NAO"])
    return out.getvalue().encode("utf-8-sig")

def table_rows(rows, limit=300):
    s = []
    for r in rows[:limit]:
        s.append("<tr>" + "".join(f"<td>{html.escape(str(v))}</td>" for v in [
            r["nf"], r["status"], fmt_date(r["data_inicial"]), fmt_date(r["data_prevista"]), fmt_date(r["data_entrega"]), r["dias_atraso"], r["parceiro"], r["filial"], r["tipo"], r["uf_origem"], r["uf_destino"], "SIM" if r["falta_link"] else "NAO"
        ]) + "</tr>")
    return "\n".join(s)

def bar_table(title, items, total):
    trs = []
    for name, val in items[:20]:
        p = pct(val, total)
        trs.append(f"<tr><td>{html.escape(str(name))}</td><td>{val}</td><td><div class='bar'><span style='width:{p}%'></span></div></td><td>{p:.2f}%</td></tr>")
    return f"<div class='card'><h3>{html.escape(title)}</h3><table><tr><th>Grupo</th><th>Qtd</th><th>Visual</th><th>%</th></tr>{''.join(trs)}</table></div>"

def sla_geral_origem_table(title, rows, total, group_key="rota_uf", group_label="origem/rota"):
    """
    Tabela completa de SLA por grupo.

    Usada para:
    - SLA Geral por Origem: group_key = "rota_uf"
    - Prazo completo por Filial: group_key = "filial"

    Regra dos indicadores por linha:
    Base do percentual = Qtd total do próprio grupo.

    SLA no prazo (%) =
        (Entregue antes do prazo + Entregue no prazo + Em aberto no prazo) / Qtd total do grupo

    Atrasado (%) =
        (Entregue atrasado + Entregue atrasado justificado + Em aberto com atraso) / Qtd total do grupo

    Aguardando confirmação de entrega fica separado e não entra nos indicadores.
    """
    agrupado = {}

    for r in rows:
        grupo = r.get(group_key) or "Nao informado"
        status = r.get("status") or "Nao informado"

        if grupo not in agrupado:
            agrupado[grupo] = {
                "Qtd": 0,
                "Entregue antes do prazo": 0,
                "Entregue no prazo": 0,
                "Entregue atrasado": 0,
                "Entregue atrasado - Justificado": 0,
                "Aguardando confirmação de entrega": 0,
                "Em aberto no prazo": 0,
                "Em aberto com atraso": 0,
            }

        agrupado[grupo]["Qtd"] += 1

        if status in agrupado[grupo]:
            agrupado[grupo][status] += 1

    ordenado = sorted(agrupado.items(), key=lambda x: x[1]["Qtd"], reverse=True)

    trs = []
    for grupo, dados in ordenado[:120]:
        qtd = dados["Qtd"]
        perc_total = pct(qtd, total)
        visual_w = min(100, perc_total)

        antes = dados["Entregue antes do prazo"]
        no_prazo = dados["Entregue no prazo"]
        atrasado = dados["Entregue atrasado"]
        atrasado_just = dados["Entregue atrasado - Justificado"]
        aguardando = dados["Aguardando confirmação de entrega"]
        aberto_prazo = dados["Em aberto no prazo"]
        aberto_atraso = dados["Em aberto com atraso"]

        # CÁLCULO CORRETO DOS INDICADORES DA LINHA:
        # Percentual calculado sobre o QTD total do próprio grupo.
        sla_ok_qtd = antes + no_prazo + aberto_prazo
        sla_atrasado_qtd = atrasado + atrasado_just + aberto_atraso

        if qtd > 0:
            sla_no_prazo_pct = round((sla_ok_qtd / qtd) * 100, 2)
            atraso_pct = round((sla_atrasado_qtd / qtd) * 100, 2)
        else:
            sla_no_prazo_pct = 0
            atraso_pct = 0

        trs.append(f"""
        <tr>
            <td class='grupo_col'>{html.escape(str(grupo))}</td>
            <td class='num_col'>{qtd}</td>
            <td class='num_col'>{perc_total:.2f}%</td>
            <td class='visual_col'><div class='bar'><span style='width:{visual_w}%'></span></div></td>
            <td class='num_col ok_txt'>{antes}</td>
            <td class='num_col prazo_txt'>{no_prazo}</td>
            <td class='num_col atraso_txt'>{atrasado}</td>
            <td class='num_col just_txt'>{atrasado_just}</td>
            <td class='num_col neutro_txt'>{aguardando}</td>
            <td class='num_col prazo_txt'>{aberto_prazo}</td>
            <td class='num_col atraso_txt'>{aberto_atraso}</td>
            <td class='num_col sla_ok_pct'>{sla_no_prazo_pct:.2f}%</td>
            <td class='num_col sla_bad_pct'>{atraso_pct:.2f}%</td>
        </tr>
        """)

    return f"""
    <div class='card sla_origem_card clean_table_card'>
        <div class='clean_head'>
            <div>
                <h3>{html.escape(title)}</h3>
                <p>Visão por {html.escape(group_label)}, com status em quantidade e indicadores em percentual.</p>
            </div>
        </div>

        <div class='tablebox sla_scroll'>
            <table class='sla_origem_table clean_sla_table'>
                <thead>
                    <tr class='group_header'>
                        <th rowspan='2'>Grupo</th>
                        <th rowspan='2'>Qtd</th>
                        <th rowspan='2'>% Total</th>
                        <th rowspan='2'>Visual</th>
                        <th colspan='4'>Entregues</th>
                        <th colspan='3'>Em aberto</th>
                        <th colspan='2'>Indicadores</th>
                    </tr>
                    <tr class='sub_header'>
                        <th>Antes do prazo</th>
                        <th>No prazo</th>
                        <th>Atrasado</th>
                        <th>Atrasado just.</th>
                        <th>Aguardando</th>
                        <th>No prazo</th>
                        <th>Com atraso</th>
                        <th>SLA no prazo</th>
                        <th>Atrasado</th>
                    </tr>
                </thead>
                <tbody>
                    {''.join(trs)}
                </tbody>
            </table>
        </div>

        <div class='sla_legenda'>
            <span><i class='dot ok'></i>Antes do prazo</span>
            <span><i class='dot prazo'></i>No prazo</span>
            <span><i class='dot atraso'></i>Atrasado / em aberto com atraso</span>
            <span><i class='dot just'></i>Atrasado justificado</span>
            <span>SLA no prazo (%) = (Antes do prazo + No prazo + Em aberto no prazo) / Qtd total do grupo</span>
        </div>
    </div>
    """

def svg_bar_chart(title, items, width=720, height=300, limit=7):
    items = list(items)
    if limit:
        items = items[:limit]
    if not items:
        return "<div class='chart_empty'>Sem dados para exibir</div>"
    maxv = max(v for _, v in items) or 1
    left = 170
    right = 58
    top = 52
    row_h = 30
    gap = 10
    height = max(height, top + len(items) * (row_h + gap) + 28)
    plot_w = width - left - right
    parts = [f"<svg class='svg_chart' viewBox='0 0 {width} {height}' role='img' preserveAspectRatio='xMidYMid meet'>"]
    parts.append(f"<text x='24' y='28' class='chart_title_left'>{html.escape(title)}</text>")
    for i in range(5):
        x = left + (plot_w / 4) * i
        parts.append(f"<line x1='{x:.1f}' y1='{top-14}' x2='{x:.1f}' y2='{height-20}' class='gridline_v'/>")
    for idx, (name, val) in enumerate(items):
        y = top + idx * (row_h + gap)
        bar_w = max(5, (val / maxv) * plot_w)
        label = html.escape(str(name))[:28]
        parts.append(f"<text x='{left-12}' y='{y+20}' text-anchor='end' class='h_label'>{label}</text>")
        parts.append(f"<rect x='{left}' y='{y}' width='{plot_w}' height='{row_h}' rx='8' class='bar_bg'/>")
        parts.append(f"<rect x='{left}' y='{y}' width='{bar_w:.1f}' height='{row_h}' rx='8' class='barblue'/>")
        parts.append(f"<text x='{left+bar_w+8:.1f}' y='{y+20}' class='valtxt'>{val}</text>")
    parts.append("</svg>")
    return ''.join(parts)

def svg_donut_chart(title, ok, bad):
    total = max(1, ok + bad)
    ok_pct = round((ok / total) * 100, 1)
    bad_pct = 100 - ok_pct
    return f"""
    <div class='donut_box'>
        <h3>{html.escape(title)}</h3>
        <div class='donut_css' style='--ok:{ok_pct};'>
            <div class='donut_center'><b>{ok_pct:.1f}%</b><span>no prazo</span></div>
        </div>
        <div class='legend_row'>
            <span><i class='legend_ok'></i>Finalizado no prazo</span>
            <span><i class='legend_bad'></i>Atrasado {bad_pct:.1f}%</span>
        </div>
    </div>
    """



def route_chart_items(rows, top_n=7):
    """
    Corrige o gráfico Top rotas UF.
    Antes o gráfico mostrava somente as 7 primeiras rotas e escondia o restante,
    dando a impressão de que a soma estava errada.
    Agora ele mostra as TOP rotas + uma barra "Outras rotas", garantindo que
    o total visualizado feche com o total geral da base.
    """
    todas = group_count(rows, 'rota_uf')
    top = todas[:top_n]
    resto = sum(v for _, v in todas[top_n:])
    if resto > 0:
        top.append(("Outras rotas", resto))
    return top

def render_charts(rows):
    s = make_summary(rows)
    status_items = sorted(s['counts'].items(), key=lambda x: x[1], reverse=True)
    no_prazo_just = s['counts'].get('Entregue antes do prazo', 0) + s['counts'].get('Entregue no prazo', 0) + s['counts'].get('Entregue atrasado - Justificado', 0)
    atrasado_just = s['counts'].get('Entregue atrasado', 0)
    rotas = route_chart_items(rows, top_n=7)
    total_rotas = sum(v for _, v in rotas)
    return f"""
    <div class='charts_grid clean'>
        <div class='chart_card status_chart'>{svg_bar_chart('Status geral', status_items, width=760, height=320)}</div>
        <div class='chart_card donut_card'>{svg_donut_chart('SLA - Justificado', no_prazo_just, atrasado_just)}</div>
        <div class='chart_card route_chart'>{svg_bar_chart('Top rotas UF + Outras | Total ' + str(total_rotas), rotas, width=980, height=370, limit=8)}</div>
    </div>
    """

def render_login(msg=""):
    return f"""<!doctype html><html lang='pt-br'><head><meta charset='utf-8'><title>Login - Dashboard SLA SIN</title>{CSS}</head><body>
    <div class='login_wrap'>
        <div class='login_card'>
            <div class='login_brand'>
                <h1>Dashboard SLA SIN</h1>
                <p>ADM Empresa + Pesquisa Cliente</p>
            </div>
            {'<div class=err>'+html.escape(msg)+'</div>' if msg else ''}
            <form method='post' action='/login' class='login_form'>
                <label>Usuário</label>
                <input type='text' name='login' autocomplete='username' required>
                <label>Senha</label>
                <input type='password' name='senha' autocomplete='current-password' required>
                <button>Entrar</button>
            </form>
            <div class='note'>Admin padrão inicial: admin / Admin@123. Altere a senha antes de publicar.</div>
        </div>
    </div></body></html>"""

def render_change_password(user, msg=""):
    return f"""<!doctype html><html lang='pt-br'><head><meta charset='utf-8'><title>Alterar senha</title>{CSS}</head><body>
    <div class='login_wrap'>
        <div class='login_card'>
            <div class='login_brand'><h1>Alterar senha</h1><p>Usuário: {html.escape(user.get('login',''))}</p></div>
            {'<div class=err>'+html.escape(msg)+'</div>' if msg else ''}
            <form method='post' action='/alterar_senha' class='login_form'>
                <label>Senha atual</label>
                <input type='password' name='senha_atual' required>
                <label>Nova senha</label>
                <input type='password' name='nova_senha' minlength='6' required>
                <label>Confirmar nova senha</label>
                <input type='password' name='confirmar_senha' minlength='6' required>
                <button>Salvar nova senha</button>
            </form>
        </div>
    </div></body></html>"""

def render_clientes_admin(rows, created_users=None):
    clientes = {}
    for r in rows:
        codigo = normalize_cliente_codigo(r.get("cliente_codigo", ""))
        if not codigo:
            continue
        if codigo not in clientes:
            clientes[codigo] = {"nome": r.get("cliente_nome", ""), "qtd": 0}
        clientes[codigo]["qtd"] += 1
    itens = sorted(clientes.items(), key=lambda x: x[1]["qtd"], reverse=True)
    trs = []
    for codigo, info in itens[:300]:
        login = cliente_login_from_codigo(codigo)
        trs.append(f"<tr><td>{html.escape(codigo)}</td><td>{html.escape(info['nome'])}</td><td>{html.escape(login)}</td><td>{info['qtd']}</td></tr>")
    novos = ""
    if created_users:
        linhas = "".join(f"<tr><td>{html.escape(a)}</td><td>{html.escape(b)}</td><td>{html.escape(c)}</td><td>{html.escape(d)}</td></tr>" for a,b,c,d in created_users[:300])
        novos = f"<div class='card'><h2>Novos usuários criados</h2><p class='sub'>Senha inicial gerada. Oriente o cliente a trocar no primeiro acesso.</p><div class='tablebox'><table><tr><th>Login</th><th>Senha inicial</th><th>Código</th><th>Cliente</th></tr>{linhas}</table></div></div>"
    return f"""<!doctype html><html lang='pt-br'><head><meta charset='utf-8'><title>Clientes</title>{CSS}</head><body>
    <div class='wrap wide'>
        <div class='top'><div><h1>Clientes identificados</h1><p>Baseada na aba Rastreamento: AF código cliente e AG nome cliente.</p></div><div class='actions'><a class='btn secondary' href='/dashboard'>Voltar</a><a class='btn secondary' href='/logout'>Sair</a></div></div>
        {novos}
        <div class='card'><h2>Relação de clientes</h2><div class='tablebox'><table><tr><th>Código Cliente</th><th>Nome Cliente</th><th>Login</th><th>Demandas</th></tr>{''.join(trs)}</table></div></div>
    </div></body></html>"""

def render_upload(msg=""):
    return f"""<!doctype html><html lang='pt-br'><head><meta charset='utf-8'><title>Dashboard SLA SIN</title>{CSS}</head><body><div class='wrap'><div class='hero'><h1>Dashboard SLA SIN</h1><p>Área ADM: suba a planilha Excel uma vez ao dia. Clientes consultam apenas suas demandas.</p></div>{'<div class=err>'+html.escape(msg)+'</div>' if msg else ''}<form class='upload' method='post' enctype='multipart/form-data' action='/analisar'><input type='file' name='file' accept='.xlsx,.xlsm' required><button>Analisar planilha</button></form><div class='note'>Colunas usadas: Rastreamento C, I, K, R, S, Y, AB, AC, AD, AF código cliente, AG nome cliente e aba de feriados.</div></div></body></html>"""

def page_link(token, kind):
    return f"/pagina?token={token}&tipo={kind}"

def resumo_card(title, value, subtitle, href, css_class="blue"):
    return f"""
    <a class='dash_btn {css_class}' href='{href}'>
        <span>{html.escape(str(title))}</span>
        <b>{html.escape(str(value))}</b>
        <small>{html.escape(str(subtitle))}</small>
    </a>
    """

def render_dashboard(rows, token, user=None):
    s = make_summary(rows)
    total = max(1, s["total"])
    atrasados = [r for r in rows if r["status"] in ("Entregue atrasado", "Entregue atrasado - Justificado")]
    aberto_atraso = [r for r in rows if r["status"] == "Em aberto com atraso"]
    aberto_prazo = [r for r in rows if r["status"] == "Em aberto no prazo"]
    justificados = [r for r in rows if r["status"] == "Entregue atrasado - Justificado"]
    falta_link = [r for r in rows if r["falta_link"]]
    devolucao = [r for r in rows if r['operacao_grupo'] == 'Devolucao/Reversa']

    status_html = "".join([
        f"<div class='stat'><span>{html.escape(k)}</span><b>{v}</b><em>{pct(v,total):.2f}%</em></div>"
        for k, v in sorted(s["counts"].items(), key=lambda x: x[1], reverse=True)
    ])

    botoes = "".join([
        resumo_card("SLA Geral por Origem", s["total"], "Analisar rotas UF", page_link(token, "prazo_rota"), "green"),
        resumo_card("Prazo completo por Filial", s["total"], "Analisar filiais", page_link(token, "prazo_filial"), "blue"),
        resumo_card("Entregas atrasadas por Origem x Destino UF", len(atrasados), "Analisar atraso por rota", page_link(token, "atraso_rota"), "orange"),
        resumo_card("Entregas atrasadas por Filial", len(atrasados), "Analisar atraso por filial", page_link(token, "atraso_filial"), "red"),
        resumo_card("Devolução/Reversa por Filial", len(devolucao), "Analisar reversas", page_link(token, "devolucao_filial"), "purple"),
        resumo_card("Relação de atraso por BASE", len(aberto_atraso), "Em aberto atrasado", page_link(token, "parceiros_atrasado"), "red"),
        resumo_card("Pedidos em aberto atrasados", len(aberto_atraso), "Lista detalhada", page_link(token, "aberto_atrasado"), "orange"),
        resumo_card("Pedidos em aberto no prazo", len(aberto_prazo), "Lista detalhada", page_link(token, "aberto_prazo"), "green"),
        resumo_card("Prazo atrasado justificado", len(justificados), "Lista detalhada", page_link(token, "justificados"), "purple"),
        resumo_card("Falta link de comprovante", len(falta_link), "Lista detalhada", page_link(token, "falta_link"), "red"),
    ])

    return f"""<!doctype html><html lang='pt-br'><head><meta charset='utf-8'><title>Dashboard SLA SIN</title>{CSS}</head><body>
    <div class='wrap wide'>
        <div class='top'>
            <div><h1>Dashboard SLA SIN</h1><p>Usuário: <b>{html.escape((user or {}).get('nome',''))}</b> | Perfil: <b>{html.escape((user or {}).get('tipo',''))}</b> | Base analisada em {datetime.now().strftime('%d/%m/%Y %H:%M')} | Total: <b>{s['total']}</b> | Consulta ativa por até {CACHE_VALID_HOURS}h</p></div>
            <div class='actions'>{"<a class='btn' href='/'>Nova análise</a><a class='btn secondary' href='/clientes'>Clientes</a>" if (user or {}).get('tipo') == 'admin' else ""}<a class='btn secondary' href='/alterar_senha'>Senha</a><a class='btn secondary' href='/logout'>Sair</a></div>
        </div>

        <section class='grid kpis'>
            <div class='kpi green'><b>{s['pct_finalizados']:.2f}%</b><span>Finalizados</span><small>{s['finalizados']} pedidos</small></div>
            <div class='kpi blue'><b>{s['pct_em_aberto']:.2f}%</b><span>Em aberto</span><small>{s['em_aberto']} pedidos</small></div>
            <div class='kpi orange'><b>{s['sla_real_ok']:.2f}%</b><span>SLA real no prazo</span><small>Atrasado real: {s['sla_real_atrasado']:.2f}%</small></div>
            <div class='kpi purple'><b>{s['sla_just_ok']:.2f}%</b><span>SLA justificado</span><small>Atrasado: {s['sla_just_atrasado']:.2f}%</small></div>
            <div class='kpi red'><b>{s['falta_link']}</b><span>Falta link</span><small>Comprovante vazio</small></div>
        </section>

        <section class='grid two'>
            <div class='card'><h2>Status geral</h2>{status_html}</div>
            <div class='card'><h2>Medidores SLA</h2>
                <div class='meter'><label>SLA Real</label><div><span style='width:{s['sla_real_ok']}%'></span></div><p>No prazo {s['sla_real_ok']:.2f}% | Atrasado {s['sla_real_atrasado']:.2f}%</p></div>
                <div class='meter'><label>SLA Justificado</label><div><span style='width:{s['sla_just_ok']}%'></span></div><p>No prazo {s['sla_just_ok']:.2f}% | Atrasado {s['sla_just_atrasado']:.2f}%</p></div>
            </div>
        </section>

        <section class='card painel_analises'>
            <h2>Análises disponíveis</h2>
            <p class='sub'>Clique em cada botão para abrir uma página separada e analisar apenas aquele ponto.</p>
            <div class='home_layout'>
                <div class='menu_grid compact'>{botoes}</div>
                {render_charts(rows)}
            </div>
        </section>
    </div></body></html>"""

def render_detail_page(rows, token, kind, user=None):
    atrasados = [r for r in rows if r["status"] in ("Entregue atrasado", "Entregue atrasado - Justificado")]
    aberto_atraso = sorted([r for r in rows if r["status"] == "Em aberto com atraso"], key=lambda r: r["dias_atraso"], reverse=True)
    aberto_prazo = sorted([r for r in rows if r["status"] == "Em aberto no prazo"], key=lambda r: r["data_prevista"] or date.max)
    justificados = [r for r in rows if r["status"] == "Entregue atrasado - Justificado"]
    falta_link = [r for r in rows if r["falta_link"]]
    devolucao = [r for r in rows if r['operacao_grupo'] == 'Devolucao/Reversa']

    config = {
        "prazo_rota": ("SLA Geral por Origem", sla_geral_origem_table('SLA Geral por Origem', rows, max(1, len(rows)), group_key='rota_uf', group_label='Origem x Destino UF'), None),
        "prazo_filial": ("Prazo completo por Filial", sla_geral_origem_table('Prazo completo por Filial', rows, max(1, len(rows)), group_key='filial', group_label='Filial'), None),
        "atraso_rota": ("Entregas atrasadas por Origem x Destino UF", bar_table('Entregas atrasadas por Origem x Destino UF', group_count(atrasados, 'rota_uf'), max(1, len(atrasados))), None),
        "atraso_filial": ("Entregas atrasadas por Filial", bar_table('Entregas atrasadas por Filial', group_count(atrasados, 'filial'), max(1, len(atrasados))), None),
        "devolucao_filial": ("Devolução/Reversa por Filial", bar_table('Devolução/Reversa por Filial', group_count(devolucao, 'filial'), max(1, len(rows))), None),
        "parceiros_atrasado": (
            "Relação de atraso por BASE",
            bar_table(
                'Relação de atraso por BASE',
                group_count(aberto_atraso, 'parceiro_uf'),
                max(1, len(aberto_atraso))
            ),
            None
        ),
        "aberto_atrasado": ("Pedidos em aberto atrasados", section_table('Pedidos em aberto atrasados', aberto_atraso, token, 'aberto_atrasado'), 'aberto_atrasado'),
        "aberto_prazo": ("Pedidos em aberto no prazo", section_table('Pedidos em aberto no prazo', aberto_prazo, token, 'aberto_prazo'), 'aberto_prazo'),
        "justificados": ("Prazo de entrega atrasado justificado", section_table('Prazo de entrega atrasado justificado', justificados, token, 'justificados'), 'justificados'),
        "falta_link": ("Falta link de comprovante", section_table('Falta link de comprovante', falta_link, token, 'falta_link'), 'falta_link'),
    }
    title, content, download_kind = config.get(kind, config["prazo_rota"])
    download = f"<a class='btn' href='/download?token={token}&tipo={download_kind}'>Baixar CSV</a>" if download_kind else ""
    return f"""<!doctype html><html lang='pt-br'><head><meta charset='utf-8'><title>{html.escape(title)}</title>{CSS}</head><body>
    <div class='wrap wide'>
        <div class='top'>
            <div><h1>{html.escape(title)}</h1><p>Análise individual do ponto selecionado.</p></div>
            <div class='actions'><a class='btn secondary' href='/dashboard'>Voltar ao resumo</a>{download}</div>
        </div>
        {content}
    </div></body></html>"""

def section_table(title, rows, token, kind):
    return f"<section class='card'><div class='section_head'><h2>{html.escape(title)} <small>({len(rows)})</small></h2><a class='btn small' href='/download?token={token}&tipo={kind}'>Baixar CSV</a></div><div class='tablebox'><table><tr><th>NF</th><th>Status</th><th>Data inicial</th><th>Data prevista</th><th>Data entrega</th><th>Dias atraso</th><th>Parceiro</th><th>Filial</th><th>Tipo</th><th>UF Origem</th><th>UF Destino</th><th>Falta link</th></tr>{table_rows(rows)}</table></div></section>"

CSS = """<style>
*{box-sizing:border-box}
body{margin:0;background:#edf3f8;font-family:Segoe UI,Arial,sans-serif;color:#17212b}
.wrap{max-width:980px;margin:26px auto;padding:0 18px}.wide{max-width:1480px}
.hero,.card,.upload{background:#fff;border-radius:18px;padding:24px;box-shadow:0 10px 28px rgba(31,63,94,.08);border:1px solid #e3edf5}
.hero{background:linear-gradient(135deg,#0b6b8f,#10b36a);color:white}.hero h1,h1,h2,h3{margin:0 0 10px}
.upload{margin-top:18px;display:flex;gap:14px;align-items:center}.upload input{flex:1;padding:15px;border:1px dashed #91a9bd;border-radius:12px;background:#f8fbfd}
button,.btn{background:#0b6b8f;color:white;border:0;border-radius:12px;padding:12px 20px;font-weight:700;text-decoration:none;display:inline-block;cursor:pointer}.btn.small{padding:8px 12px;font-size:13px}.btn.secondary{background:#607d8b}
.note{margin-top:15px;color:#587083}.err{background:#ffe8e8;color:#8d1f1f;border:1px solid #ffc4c4;padding:14px;border-radius:12px;margin-top:15px}
.top{display:flex;justify-content:space-between;align-items:center;margin-bottom:18px}.top p{margin:0;color:#516879}.actions{display:flex;gap:10px;align-items:center}
.grid{display:grid;gap:14px;margin-bottom:14px}.kpis{grid-template-columns:repeat(5,1fr)}.two{grid-template-columns:1fr 1fr}
.kpi{border-radius:16px;padding:16px;color:white;min-height:104px;box-shadow:0 8px 22px rgba(31,63,94,.10)}.kpi b{font-size:28px;display:block}.kpi span{display:block;font-size:14px;font-weight:700}.kpi small{opacity:.9}
.green{background:#00a957}.blue{background:#159ad1}.orange{background:#f06b2f}.purple{background:#8257d8}.red{background:#c94343}
.stat{display:grid;grid-template-columns:1fr 90px 90px;gap:10px;border-bottom:1px solid #eef3f7;padding:10px 0}.stat b{text-align:right}.stat em{text-align:right;color:#62788b;font-style:normal}
.meter{margin:14px 0}.meter label{font-weight:800}.meter>div{height:18px;background:#edf2f6;border-radius:20px;overflow:hidden;margin-top:6px}.meter span{height:100%;display:block;background:#00a957}
.bar{height:13px;background:#edf2f6;border-radius:20px;overflow:hidden;min-width:100px}.bar span{height:100%;display:block;background:#0b6b8f}
table{border-collapse:collapse;width:100%;font-size:13px}th,td{border-bottom:1px solid #e8eef4;text-align:left;padding:8px 10px;white-space:nowrap}th{background:#e7f3ec;color:#102b1d;font-weight:800;position:sticky;top:0}tr:hover td{background:#f7fbff}
.tablebox{overflow:auto;max-height:520px}.section_head{display:flex;justify-content:space-between;align-items:center;margin-bottom:8px}section.card{margin-bottom:14px}.sub{color:#587083;margin-top:0}
.painel_analises{padding:24px}.painel_analises h2{font-size:22px}.home_layout{display:grid;grid-template-columns:390px 1fr;gap:18px;align-items:start}
.menu_grid.compact{display:grid;grid-template-columns:repeat(2,1fr);gap:9px}.dash_btn{display:block;text-decoration:none;color:white;border-radius:14px;padding:11px 12px;min-height:82px;box-shadow:0 5px 16px rgba(31,63,94,.11);transition:.15s;overflow:hidden}.dash_btn:hover{transform:translateY(-2px);filter:brightness(1.04)}.dash_btn span{display:block;font-size:12px;line-height:1.15;font-weight:800}.dash_btn b{display:block;font-size:24px;margin-top:8px;letter-spacing:.2px}.dash_btn small{display:block;margin-top:3px;font-size:11px;opacity:.95}
.charts_grid.clean{display:grid;grid-template-columns:1.25fr .85fr;gap:14px}.chart_card{background:#fff;border:1px solid #dfe8f0;border-radius:16px;padding:14px;min-height:300px;box-shadow:0 7px 20px rgba(31,63,94,.06);overflow:hidden}.route_chart{grid-column:1 / -1;min-height:300px}.svg_chart{width:100%;height:100%;display:block}.chart_title_left{font-size:18px;font-weight:800;fill:#344b5a}.gridline_v{stroke:#edf2f6;stroke-width:1}.bar_bg{fill:#edf3f7}.barblue{fill:#0f7897}.valtxt{font-size:13px;fill:#344b5a;font-weight:800}.h_label{font-size:13px;fill:#334c5b;font-weight:600}.chart_empty{padding:30px;text-align:center;color:#607d8b}
.donut_card{display:flex;align-items:center;justify-content:center}.donut_box{text-align:center;width:100%}.donut_box h3{font-size:18px;color:#344b5a;margin-bottom:18px}.donut_css{--size:190px;width:var(--size);height:var(--size);border-radius:50%;margin:8px auto 18px;background:conic-gradient(#0f7897 calc(var(--ok)*1%), #f3702f 0);display:flex;align-items:center;justify-content:center;box-shadow:inset 0 0 0 1px rgba(0,0,0,.03)}.donut_center{width:112px;height:112px;background:white;border-radius:50%;display:flex;flex-direction:column;align-items:center;justify-content:center;box-shadow:0 2px 10px rgba(31,63,94,.08)}.donut_center b{font-size:28px;color:#0f7897;line-height:1}.donut_center span{font-size:12px;color:#607d8b;margin-top:7px}.legend_row{display:flex;gap:14px;justify-content:center;flex-wrap:wrap;font-size:12px;color:#394b59}.legend_row i{width:10px;height:10px;display:inline-block;border-radius:2px;margin-right:6px}.legend_ok{background:#0f7897}.legend_bad{background:#f3702f}

.clean_table_card{
    padding:22px 22px 18px;
    background:#fff;
    border:2px solid #b7dcc8;
    border-radius:16px;
    box-shadow:0 8px 22px rgba(31,63,94,.07)
}
.clean_head{
    display:flex;
    align-items:flex-start;
    justify-content:space-between;
    margin-bottom:14px
}
.clean_head h3{
    font-size:23px;
    color:#0b1f36;
    margin:0 0 6px;
    font-weight:600
}
.clean_head p{
    margin:0;
    color:#587083;
    font-size:14px;
    font-weight:400
}
.sla_scroll{
    max-height:690px;
    border:1px solid #cfe8db;
    border-radius:12px;
    overflow:auto;
    background:white
}
.clean_sla_table{
    min-width:1320px;
    font-size:13px;
    border-collapse:separate;
    border-spacing:0
}
.clean_sla_table th{
    background:#e7f3ec;
    color:#10233d;
    border-bottom:1px solid #cfe8db;
    border-right:1px solid #d9ebe2;
    text-align:center;
    padding:12px 10px;
    font-weight:500
}
.clean_sla_table .group_header th{
    font-size:13px;
    background:#dff0e7;
    font-weight:500
}
.clean_sla_table .sub_header th{
    font-size:12px;
    background:#edf7f2;
    font-weight:500
}
.clean_sla_table td{
    padding:11px 10px;
    text-align:center;
    border-bottom:1px solid #e7eef5;
    border-right:1px solid #eef3f7;
    color:#10233d;
    font-weight:400
}
.clean_sla_table tbody tr:nth-child(odd) td{
    background:#ffffff
}
.clean_sla_table tbody tr:nth-child(even) td{
    background:#f3f9fd
}
.clean_sla_table tbody tr:hover td{
    background:#eaf6ff
}
.clean_sla_table .grupo_col{
    text-align:left;
    min-width:130px;
    font-weight:400
}
.clean_sla_table .num_col{
    font-weight:400
}
.clean_sla_table .visual_col{
    min-width:180px
}
.clean_sla_table .bar{
    height:12px;
    background:#eaf1f5;
    border-radius:20px;
    overflow:hidden;
    min-width:160px
}
.clean_sla_table .bar span{
    height:100%;
    display:block;
    background:#0b6b8f;
    border-radius:20px
}
.ok_txt{color:#0b8d38!important}
.prazo_txt{color:#1268c4!important}
.atraso_txt{color:#f05a1a!important}
.just_txt{color:#a20db5!important}
.neutro_txt{color:#374151!important}
.sla_ok_pct{color:#0b8d38!important}
.sla_bad_pct{color:#f05a1a!important}
.sla_legenda{
    display:flex;
    flex-wrap:wrap;
    gap:14px;
    margin-top:14px;
    color:#587083;
    font-size:12px;
    font-weight:400
}
.sla_legenda span{
    display:inline-flex;
    align-items:center;
    gap:6px;
    font-weight:400
}
.dot{
    width:10px;
    height:10px;
    border-radius:50%;
    display:inline-block
}
.dot.ok{background:#0b8d38}
.dot.prazo{background:#1268c4}
.dot.atraso{background:#f05a1a}
.dot.just{background:#a20db5}


.login_wrap{min-height:100vh;display:flex;align-items:center;justify-content:center;padding:24px;background:linear-gradient(135deg,#eaf4fb,#f3fbf6)}
.login_card{width:100%;max-width:430px;background:#fff;border-radius:20px;padding:28px;box-shadow:0 14px 34px rgba(31,63,94,.14);border:1px solid #e3edf5}
.login_brand{background:linear-gradient(135deg,#0b6b8f,#10b36a);color:#fff;border-radius:16px;padding:20px;margin-bottom:18px}
.login_brand h1{font-size:25px;margin-bottom:6px}.login_brand p{margin:0;opacity:.95}
.login_form{display:flex;flex-direction:column;gap:9px}
.login_form label{font-weight:700;color:#344b5a;font-size:13px}
.login_form input{padding:13px 14px;border:1px solid #cbdce8;border-radius:12px;font-size:15px}
.login_form button{margin-top:8px}
@media(max-width:1200px){.home_layout{grid-template-columns:1fr}.charts_grid.clean{grid-template-columns:1fr 1fr}}
@media(max-width:1000px){.kpis,.two,.menu_grid.compact,.charts_grid.clean{grid-template-columns:1fr}.upload,.top{display:block}.btn{margin-top:12px}.actions{display:block}.home_layout{grid-template-columns:1fr}}
</style>"""


class App(BaseHTTPRequestHandler):
    def send_html(self, body, code=200, extra_headers=None):
        data = body.encode("utf-8")
        self.send_response(code)
        if extra_headers:
            for k, v in extra_headers:
                self.send_header(k, v)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def redirect(self, location, extra_headers=None):
        self.send_response(302)
        if extra_headers:
            for k, v in extra_headers:
                self.send_header(k, v)
        self.send_header("Location", location)
        self.end_headers()

    def read_form_urlencoded(self):
        length = int(self.headers.get("Content-Length", "0"))
        raw = self.rfile.read(length).decode("utf-8", "ignore")
        data = parse_qs(raw)
        return {k: v[0] if v else "" for k, v in data.items()}

    def current_rows_and_token(self, user):
        token = load_latest_token()
        rows = load_cache(token) if token else None
        if not rows:
            return None, token
        return filter_rows_for_user(rows, user), token

    def do_GET(self):
        parsed = urlparse(self.path)

        if parsed.path == "/login":
            self.send_html(render_login())
            return

        if parsed.path == "/logout":
            self.send_response(302)
            clear_session(self)
            self.send_header("Location", "/login")
            self.end_headers()
            return

        user = get_session_user(self)
        if not user:
            self.send_html(render_login("Faça login para acessar."), 401)
            return

        if user.get("trocar_senha") and parsed.path not in ("/alterar_senha",):
            self.send_html(render_change_password(user, "Por segurança, altere sua senha inicial."))
            return

        if parsed.path == "/alterar_senha":
            self.send_html(render_change_password(user))
            return

        if parsed.path == "/":
            if user.get("tipo") != "admin":
                self.redirect("/dashboard")
                return
            self.send_html(render_upload())
            return

        if parsed.path == "/clientes":
            if user.get("tipo") != "admin":
                self.send_html(render_login("Acesso restrito ao administrador."), 403)
                return
            rows, token = self.current_rows_and_token(user)
            if rows is None:
                self.send_html(render_upload("Nenhuma análise carregada ainda. Suba a planilha."))
                return
            self.send_html(render_clientes_admin(rows))
            return

        if parsed.path == "/dashboard":
            qs = parse_qs(parsed.query)
            token = qs.get("token", [""])[0] or load_latest_token()
            rows = load_cache(token) if token else None
            if not rows:
                if user.get("tipo") == "admin":
                    self.send_html(render_upload("Analise nao encontrada ou vencida. Suba a planilha novamente."), 404)
                else:
                    self.send_html(render_login("Nenhuma pesquisa disponível para seu cliente no momento."), 404)
                return
            rows_user = filter_rows_for_user(rows, user)
            self.send_html(render_dashboard(rows_user, token, user))
            return

        if parsed.path == "/pagina":
            qs = parse_qs(parsed.query)
            token = qs.get("token", [""])[0] or load_latest_token()
            tipo = qs.get("tipo", ["prazo_rota"])[0]
            rows = load_cache(token) if token else None
            if not rows:
                self.send_html(render_login("Analise nao encontrada ou vencida."), 404)
                return
            rows_user = filter_rows_for_user(rows, user)
            self.send_html(render_detail_page(rows_user, token, tipo, user))
            return

        if parsed.path == "/download":
            qs = parse_qs(parsed.query)
            token = qs.get("token", [""])[0] or load_latest_token()
            tipo = qs.get("tipo", ["todos"])[0]
            rows = load_cache(token) if token else None
            if not rows:
                self.send_html(render_login("Analise nao encontrada ou vencida."), 404)
                return
            rows = filter_rows_for_user(rows, user)
            filters = {
                "aberto_atrasado": lambda r: r["status"] == "Em aberto com atraso",
                "aberto_prazo": lambda r: r["status"] == "Em aberto no prazo",
                "justificados": lambda r: r["status"] == "Entregue atrasado - Justificado",
                "falta_link": lambda r: r["falta_link"],
                "todos": lambda r: True,
            }
            selected = [r for r in rows if filters.get(tipo, filters["todos"])(r)]
            data = csv_bytes(selected)
            self.send_response(200)
            self.send_header("Content-Type", "text/csv; charset=utf-8")
            self.send_header("Content-Disposition", f"attachment; filename={tipo}_sla_sin.csv")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return

        self.send_html(render_login("Pagina nao encontrada."), 404)

    def do_POST(self):
        parsed = urlparse(self.path)

        if parsed.path == "/login":
            form = self.read_form_urlencoded()
            login = norm_text(form.get("login", ""))
            senha = form.get("senha", "")
            users = load_users()
            u = users.get(login)
            if not u or not check_password(senha, u.get("senha_hash", "")):
                self.send_html(render_login("Usuário ou senha inválidos."), 401)
                return
            self.send_response(302)
            create_session(self, login)
            self.send_header("Location", "/dashboard" if u.get("tipo") != "admin" else "/")
            self.end_headers()
            return

        user = get_session_user(self)
        if not user:
            self.send_html(render_login("Faça login para acessar."), 401)
            return

        if parsed.path == "/alterar_senha":
            form = self.read_form_urlencoded()
            senha_atual = form.get("senha_atual", "")
            nova = form.get("nova_senha", "")
            confirmar = form.get("confirmar_senha", "")
            users = load_users()
            stored = users.get(user["login"], {})
            if not check_password(senha_atual, stored.get("senha_hash", "")):
                self.send_html(render_change_password(user, "Senha atual incorreta."), 400)
                return
            if len(nova) < 6:
                self.send_html(render_change_password(user, "A nova senha precisa ter pelo menos 6 caracteres."), 400)
                return
            if nova != confirmar:
                self.send_html(render_change_password(user, "A confirmação não confere."), 400)
                return
            users[user["login"]]["senha_hash"] = hash_password(nova)
            users[user["login"]]["trocar_senha"] = False
            save_users(users)
            self.redirect("/dashboard" if user.get("tipo") != "admin" else "/")
            return

        if parsed.path != "/analisar":
            self.send_html(render_login("Rota invalida."), 404)
            return

        if user.get("tipo") != "admin":
            self.send_html(render_login("Somente o administrador pode subir planilha."), 403)
            return

        try:
            ctype = self.headers.get("Content-Type", "")
            m = re.search("boundary=(.*)", ctype)
            if not m:
                raise ValueError("Upload invalido.")
            boundary = ("--" + m.group(1)).encode()
            length = int(self.headers.get("Content-Length", "0"))
            body = self.rfile.read(length)
            parts = body.split(boundary)
            file_bytes = None
            filename = "upload.xlsx"
            for part in parts:
                if b'name="file"' in part and b'filename=' in part:
                    head, content = part.split(b"\r\n\r\n", 1)
                    content = content.rsplit(b"\r\n", 1)[0]
                    file_bytes = content
                    fm = re.search(rb'filename="([^"]+)"', head)
                    if fm:
                        filename = fm.group(1).decode("utf-8", "ignore") or filename
                    break
            if not file_bytes:
                raise ValueError("Nenhum arquivo recebido.")
            safe = re.sub(r"[^A-Za-z0-9_.-]", "_", filename)
            path = os.path.join(UPLOAD_DIR, f"{uuid.uuid4().hex}_{safe}")
            with open(path, "wb") as f:
                f.write(file_bytes)
            rows, holidays = read_workbook(path)
            token = uuid.uuid4().hex
            CACHE[token] = rows
            save_cache(token, rows)
            save_latest_token(token)
            created = auto_create_client_users(rows)
            if created:
                self.send_html(render_clientes_admin(rows, created))
            else:
                self.send_html(render_dashboard(rows, token, user))
        except Exception as e:
            traceback.print_exc()
            self.send_html(render_upload("Erro ao analisar: " + str(e)), 500)

if __name__ == "__main__":
    load_users()
    print("Dashboard SLA SIN iniciado")
    print(f"Abra no navegador: http://{HOST}:{PORT}")
    print(f"Login admin inicial: {ADMIN_LOGIN}")
    print("Senha admin inicial: definida por ADMIN_PASSWORD ou Admin@123")
    HTTPServer((HOST, PORT), App).serve_forever()
